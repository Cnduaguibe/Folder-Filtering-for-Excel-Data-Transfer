# Objective: Read ORU data from text files and filter transmittance values, then average values for each file and print
import os
import re
import openpyxl as pyxl

# This variable asks the user for the file path containing the folder where data files are stored.
print()
#path = input("Enter ORU folder file path: ")
path = r"C:\Users\cnduaguibe\Desktop\RunTest3\Test 3b"
testname = path[37:45]
#savefile = input("Enter Template Name: ")
savefile = r"C:\Users\cnduaguibe\Desktop\RunTest3(500AG).xlsx"

def copy_readings_from_file(sheet, readings_list, row, col):
    for i in range(25):
        c = sheet.cell(row = row, column = col)
        readings_list.append(c.value)
        row += 1

def write_readings_to_template(sheet, readings_list, row, col):
    for i in range(25):
        c = sheet.cell(row = row, column = col)
        c.value = (readings_list[i])
        row += 1


count = 0

#filename = input("Enter File Save Name: ")
filename = "RunTest3(500AG).xlsx"
path2 = r"C:\Users\cnduaguibe\Desktop"
filesave_path = os.path.join(path2, filename)
template = pyxl.load_workbook(savefile)
template_sheet = template[testname]

for file in os.listdir(path):
    with open(os.path.join(path, file)) as opened_file:
        if file.endswith(".xlsx"):
            ch1_readings = []
            ch2_readings = []
            ch3_readings = []
            ch4_readings = []
            print(file)
            count += 1
            print(count)
            filepath = os.path.join(path, file)
            workbook_open = pyxl.load_workbook(filepath)
            Sheet = workbook_open["Raw Data"]
            copy_readings_from_file(Sheet, ch1_readings, 2, 2)
            copy_readings_from_file(Sheet, ch2_readings, 2, 4)
            copy_readings_from_file(Sheet, ch3_readings, 2, 6)
            copy_readings_from_file(Sheet, ch4_readings, 2, 8)
            workbook_open.close()
            print(ch1_readings)
            if count == 1:
                write_readings_to_template(template_sheet, ch1_readings, 12, 10)
                write_readings_to_template(template_sheet, ch2_readings, 12, 11)
                write_readings_to_template(template_sheet, ch3_readings, 12, 12)
                write_readings_to_template(template_sheet, ch4_readings, 12, 13)
                template.save(filesave_path)
            elif count == 2:
                write_readings_to_template(template_sheet, ch1_readings, 47, 10)
                write_readings_to_template(template_sheet, ch2_readings, 47, 11)
                write_readings_to_template(template_sheet, ch3_readings, 47, 12)
                write_readings_to_template(template_sheet, ch4_readings, 47, 13)
                template.save(filesave_path)
            elif count == 3:
                write_readings_to_template(template_sheet, ch1_readings, 12, 22)
                write_readings_to_template(template_sheet, ch2_readings, 12, 23)
                write_readings_to_template(template_sheet, ch3_readings, 12, 24)
                write_readings_to_template(template_sheet, ch4_readings, 12, 25)
                template.save(filesave_path)
            elif count == 4:
                write_readings_to_template(template_sheet, ch1_readings, 47, 22)
                write_readings_to_template(template_sheet, ch2_readings, 47, 23)
                write_readings_to_template(template_sheet, ch3_readings, 47, 24)
                write_readings_to_template(template_sheet, ch4_readings, 47, 25)
                template.save(filesave_path)
            elif count == 5:
                write_readings_to_template(template_sheet, ch1_readings, 12, 34)
                write_readings_to_template(template_sheet, ch2_readings, 12, 35)
                write_readings_to_template(template_sheet, ch3_readings, 12, 36)
                write_readings_to_template(template_sheet, ch4_readings, 12, 37)
                template.save(filesave_path)
            elif count == 6:
                write_readings_to_template(template_sheet, ch1_readings, 47, 34)
                write_readings_to_template(template_sheet, ch2_readings, 47, 35)
                write_readings_to_template(template_sheet, ch3_readings, 47, 36)
                write_readings_to_template(template_sheet, ch4_readings, 47, 37)
                template.save(filesave_path)
            elif count == 7:
                write_readings_to_template(template_sheet, ch1_readings, 12, 46)
                write_readings_to_template(template_sheet, ch2_readings, 12, 47)
                write_readings_to_template(template_sheet, ch3_readings, 12, 48)
                write_readings_to_template(template_sheet, ch4_readings, 12, 49)
                template.save(filesave_path)
            elif count == 8:
                write_readings_to_template(template_sheet, ch1_readings, 47, 46)
                write_readings_to_template(template_sheet, ch2_readings, 47, 47)
                write_readings_to_template(template_sheet, ch3_readings, 47, 48)
                write_readings_to_template(template_sheet, ch4_readings, 47, 49)
                template.save(filesave_path)
        else:
            continue



